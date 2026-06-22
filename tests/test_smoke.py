"""Smoke tests: garantem que o pacote importa e o pipeline está conectado."""

import io
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

import openpyxl
from streamlit.testing.v1 import AppTest

from before_after import __version__
from before_after.context import build_context
from before_after.correspondence import prepare
from before_after.filenames import comparative_filename
from before_after.loaders import load_correspondence, load_table
from before_after.models import METRICS, evolucao
from before_after.pairing import build_pairs
from before_after.rendering import render

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
TEMPLATE = ROOT / "templates" / "comparativo.xlsx"


def _pipeline():
    mapping = load_correspondence(str(DATA / "correspondencia.json"))
    ini = prepare(load_table(str(DATA / "dados_inicio.xlsx")), mapping)
    fim = prepare(load_table(str(DATA / "dados_final.xlsx")), mapping)
    return build_pairs(ini.frame, fim.frame)


def test_version():
    assert __version__


def test_nome_do_arquivo_inclui_paciente_e_remove_caracteres_invalidos():
    assert comparative_filename("  Maria da Conceição  ") == (
        "Quadro comparativo - Maria da Conceição.xlsx"
    )
    assert comparative_filename('Ana / Paula: "Teste"') == (
        "Quadro comparativo - Ana - Paula - Teste.xlsx"
    )
    assert comparative_filename("   ") == "Quadro comparativo.xlsx"


def test_app_exibe_campo_de_nome_do_paciente():
    app = AppTest.from_file(str(ROOT / "app.py")).run()

    assert not list(app.exception)
    assert app.text_input[0].label == "Nome do paciente"
    app.text_input[0].set_value("Maria da Conceição").run()
    assert app.text_input[0].value == "Maria da Conceição"


def test_pipeline_pareia_inicio_e_final():
    mapping = load_correspondence(str(DATA / "correspondencia.json"))
    ini = prepare(load_table(str(DATA / "dados_inicio.xlsx")), mapping)
    fim = prepare(load_table(str(DATA / "dados_final.xlsx")), mapping)

    result = build_pairs(ini.frame, fim.frame)

    assert result.pairs, "deveria produzir pares"
    # estrutura de cada par
    p = result.pairs[0]
    assert set(p.inicio) == set(METRICS)
    assert set(p.final) == set(METRICS)


def test_correspondencia_um_para_n_gera_sinonimo():
    # ADRENALINA (1 linha) -> Adrenalina + Noradrenalina (sinônimos).
    mapping = load_correspondence(str(DATA / "correspondencia.json"))
    fim = prepare(load_table(str(DATA / "dados_final.xlsx")), mapping)
    assert "Noradrenalina" in fim.frame.index


def test_evolucao_shape_invertida():
    # KOD/E-level crescem (final-início); Shape é invertida (início-final).
    assert evolucao("KOD", 1.0, 1.5) == 0.5
    assert evolucao("Shape", 4, 3) == 1
    assert evolucao("Shape", 4, 5) == -1


def test_render_preenche_template_e_espelha_brancos():
    context = build_context(_pipeline())
    out = render(str(TEMPLATE), context)

    wb = openpyxl.load_workbook(io.BytesIO(out.data))
    assert wb.sheetnames == ["Comparativo"], "aba oculta deve ser removida"
    ws = wb.active

    # Linha de dado preenchida (Hipófise, KOD na coluna C).
    assert ws.cell(3, 2).value == "Hipófise"
    assert isinstance(ws.cell(3, 3).value, (int, float))

    # Aplicabilidade: Medula óssea sem KOD (r81) fica vazia; com KOD (r96) preenche.
    assert ws.cell(81, 2).value == "Medula óssea" and ws.cell(81, 3).value is None
    assert ws.cell(96, 2).value == "Medula óssea" and ws.cell(96, 3).value is not None


def test_render_copia_datas_da_primeira_coluna_das_tabelas():
    table_ini = load_table(str(DATA / "dados_inicio.xlsx"))
    table_fim = load_table(str(DATA / "dados_final.xlsx"))
    table_ini.iat[0, 0] = datetime(2031, 1, 2)
    table_fim.iat[0, 0] = datetime(2031, 3, 4)
    period_values = (table_ini.iat[0, 0], table_fim.iat[0, 0])

    out = render(str(TEMPLATE), build_context(_pipeline()), period_values=period_values)
    ws = openpyxl.load_workbook(io.BytesIO(out.data)).active

    for col in (3, 6, 9):
        assert ws.cell(2, col).value == period_values[0]
        assert ws.cell(2, col + 1).value == period_values[1]


def test_template_e_render_nao_contem_referencias_externas():
    context = build_context(_pipeline())
    out = render(str(TEMPLATE), context)

    for workbook in (TEMPLATE, io.BytesIO(out.data)):
        with ZipFile(workbook) as archive:
            assert not any(
                name.startswith("xl/externalLinks/") for name in archive.namelist()
            )


def test_render_filtra_bloco_reprodutor_por_sistema():
    context = build_context(_pipeline())
    out = render(str(TEMPLATE), context, keep_systems={"Reprodutor Masculino"})

    wb = openpyxl.load_workbook(io.BytesIO(out.data))
    ws = wb.active
    sistemas = {ws.cell(row, 1).value for row in range(1, ws.max_row + 1)}

    assert "Reprodutor Masculino" in sistemas
    assert "Reprodutor Feminino" not in sistemas
    assert "Pénis" in {ws.cell(row, 2).value for row in range(1, ws.max_row + 1)}
    assert "Colo de útero" not in {ws.cell(row, 2).value for row in range(1, ws.max_row + 1)}
