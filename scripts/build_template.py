"""Gera o template `templates/comparativo.xlsx` a partir da aba Comparativo.

A aba "Comparativo" da planilha-fonte da equipe é a verdade de referência do
layout. Este script a transforma em um template reutilizável:

- remove a 1ª coluna ("Item Oberon", o nome cru usado como chave das fórmulas);
- troca as fórmulas SUMIFS/evolução por **células vazias** (o pipeline preenche
  os valores depois, via `rendering.render`);
- preserva cabeçalho, agrupamento por Sistema (células mescladas), estilos e
  formatos numéricos.

A planilha-fonte não é versionada (fica em `tmp/`), mas o template gerado sim.
Rode quando o layout de referência mudar:

    uv run python scripts/build_template.py [origem.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = ROOT / "tmp" / "conteudo_basico.xlsx"
OUTPUT = ROOT / "templates" / "comparativo.xlsx"
SHEET = "Comparativo"
# Aba oculta com o mapa de aplicabilidade (quais métricas a equipe preenche por
# linha). O renderer a lê e a remove antes de gerar o documento final.
MASK_SHEET = "_aplicabilidade"

# Layout da aba-fonte (antes de remover a coluna A).
HEADER_ROWS = 2
FIRST_DATA_ROW = 3
# Datas dos períodos (linha 2), lidas das células KOD da fonte.
DATE_CELLS_SOURCE = ("D2", "E2")  # início, final
# Coluna do nome do item e da célula "início" de cada métrica, na fonte.
SRC_ITEM_COL = 3
SRC_METRIC_INI_COL = {"KOD": 4, "E-level": 7, "Shape": 10}

# Colunas de evolução no template (após remover a col A): KOD, E-level, Shape.
EVOLUTION_COLS = (5, 8, 11)
# Última coluna útil (Evolução Shape); o que vier à direita é lixo da fonte.
LAST_COL = 11
# Cores da evolução: verde (>0), vermelho (<0), branco (=0). ARGB opaco em
# start_color E end_color — a formatação condicional usa o bgColor (end_color).
FILL_POS = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # verde
FILL_NEG = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # vermelho
FILL_ZERO = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # branco


def build(source: Path = DEFAULT_SOURCE, output: Path = OUTPUT) -> Path:
    wb = openpyxl.load_workbook(source)
    ws = wb[SHEET]

    # Mantém só a aba Comparativo (o template é independente da Original).
    for name in list(wb.sheetnames):
        if name != SHEET:
            del wb[name]

    inicio = ws[DATE_CELLS_SOURCE[0]].value
    final = ws[DATE_CELLS_SOURCE[1]].value

    last_data_row = max(
        (r for r in range(FIRST_DATA_ROW, ws.max_row + 1) if ws.cell(r, 3).value is not None),
        default=HEADER_ROWS,
    )

    # Aplicabilidade por linha: uma métrica é "preenchível" se a fonte tinha
    # fórmula/valor na sua célula de início (a equipe deixa KOD vazio em alguns
    # itens de propósito). Capturado por linha pois o mesmo nome de item pode ter
    # aplicabilidades diferentes (ex.: "Medula óssea" aparece com e sem KOD).
    mask = {
        row: [
            metric
            for metric, col in SRC_METRIC_INI_COL.items()
            if ws.cell(row, col).value is not None
        ]
        for row in range(FIRST_DATA_ROW, last_data_row + 1)
        if ws.cell(row, SRC_ITEM_COL).value is not None
    }

    # Guarda as mesclagens para recriá-las deslocadas (delete_cols não as ajusta).
    old_merges = [(m.min_col, m.min_row, m.max_col, m.max_row) for m in ws.merged_cells.ranges]

    ws.delete_cols(1)  # remove a coluna "Item Oberon"; valores/estilos deslocam, mesclagens não.

    # Descarta as mesclagens (agora desalinhadas) direto na coleção: `unmerge_cells`
    # apagaria valores já deslocados que caem sob uma mesclagem estale (ex.: E1).
    ws.merged_cells.ranges.clear()
    for min_col, min_row, max_col, max_row in old_merges:
        if max_col < 2:  # mesclagem que vivia só na coluna removida
            continue
        ws.merge_cells(
            start_column=min_col - 1, start_row=min_row,
            end_column=max_col - 1, end_row=max_row,
        )

    # Remove colunas à direita de "Evolução Shape" (notas/realces da equipe que
    # sobraram da fonte, ex.: realce amarelo na coluna de anotações).
    if ws.max_column > LAST_COL:
        ws.delete_cols(LAST_COL + 1, ws.max_column - LAST_COL)

    # Colunas após remover A: A=Sistema, B=Item, C..K = métricas (3 grupos de 3).
    # Datas dos períodos sob cada par início/final (KOD C/D, E-level F/G, Shape I/J).
    for ini_col, fim_col in ((3, 4), (6, 7), (9, 10)):
        ws.cell(HEADER_ROWS, ini_col, inicio)
        ws.cell(HEADER_ROWS, fim_col, final)

    # Esvazia as células numéricas (mantém formato); o renderer preenche depois.
    for row in range(FIRST_DATA_ROW, last_data_row + 1):
        for col in range(3, 12):
            ws.cell(row, col).value = None

    # Formatação condicional das colunas de evolução: verde/vermelho/branco.
    # Zera a CF herdada da fonte (desalinhada pelo delete_cols) antes de aplicar.
    ws.conditional_formatting = ConditionalFormattingList()
    for col in EVOLUTION_COLS:
        letter = get_column_letter(col)
        rng = f"{letter}{FIRST_DATA_ROW}:{letter}{last_data_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=FILL_POS))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=FILL_NEG))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=FILL_ZERO))

    # Aba oculta com a aplicabilidade: linha | KOD | E-level | Shape (1/0).
    mask_ws = wb.create_sheet(MASK_SHEET)
    mask_ws.sheet_state = "hidden"
    mask_ws.append(["row", *SRC_METRIC_INI_COL])
    for row, metrics in mask.items():
        mask_ws.append([row, *(int(m in metrics) for m in SRC_METRIC_INI_COL)])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    out = build(src)
    print(f"Template gerado: {out.relative_to(ROOT)}")
