"""Preenche o template .xlsx (formato Comparativo) com o contexto do pipeline.

O template (`templates/comparativo.xlsx`, gerado por `scripts/build_template.py`)
traz a estrutura pronta — cabeçalho, agrupamento por Sistema e a coluna `Item` —
com as células numéricas vazias. Aqui casamos cada linha pelo nome do item e
escrevemos início/final/evolução de cada métrica, preservando estilo e formato.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import IO

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

from .models import METRICS

# Layout do template (1-based). A=Sistema, B=Item, depois 3 colunas por métrica.
SYSTEM_COL = 1
ITEM_COL = 2
FIRST_DATA_ROW = 3
# Coluna inicial (início) de cada métrica; final = +1, evolução = +2.
METRIC_START_COL = {"KOD": 3, "E-level": 6, "Shape": 9}
# Aba oculta com a aplicabilidade por linha (ver scripts/build_template.py).
MASK_SHEET = "_aplicabilidade"
REPRODUCTIVE_SYSTEMS = {"Reprodutor Masculino", "Reprodutor Feminino"}


def _load_mask(wb) -> dict[int, set[str]] | None:
    """Lê a aba de aplicabilidade: linha do template -> métricas preenchíveis."""
    if MASK_SHEET not in wb.sheetnames:
        return None
    ws = wb[MASK_SHEET]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}
    metrics = header[1:]
    return {
        int(row[0]): {m for m, flag in zip(metrics, row[1:]) if flag}
        for row in rows
        if row[0] is not None
    }


def _row_systems(ws) -> dict[int, str | None]:
    """Mapeia cada linha ao Sistema, propagando células mescladas para baixo."""
    systems: dict[int, str | None] = {}
    current = None
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        value = ws.cell(row, SYSTEM_COL).value
        if value is not None:
            current = str(value).strip()
        systems[row] = current
    return systems


def _delete_rows_preserving_merges(ws, rows: list[int]) -> None:
    """Remove linhas e realinha ranges mesclados abaixo delas."""
    if not rows:
        return

    merged_ranges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
    for merged_range in merged_ranges:
        ws.unmerge_cells(merged_range)

    for row in sorted(rows, reverse=True):
        ws.delete_rows(row)

    deleted = set(rows)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merged_range)
        if any(row in deleted for row in range(min_row, max_row + 1)):
            continue

        rows_before = sum(1 for row in deleted if row < min_row)
        min_row -= rows_before
        max_row -= rows_before
        ws.merge_cells(
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )


@dataclass
class RenderResult:
    """Bytes do .xlsx + relatório de itens que não casaram com o template."""

    data: bytes
    sem_dados: list[str] = field(default_factory=list)  # linhas do template sem item no contexto
    nao_colocados: list[str] = field(default_factory=list)  # itens do contexto ausentes no template


def render(
    template: str | IO[bytes],
    context: dict[str, dict[str, dict[str, float]]],
    *,
    keep_systems: set[str] | None = None,
) -> RenderResult:
    """Preenche o template com o contexto e devolve os bytes do .xlsx.

    `template` aceita caminho ou file-like. Itens com nome repetido no template
    (ex.: "Medula óssea", "Bexiga") recebem o mesmo valor em todas as linhas.
    """
    wb = openpyxl.load_workbook(template)
    mask = _load_mask(wb)
    if MASK_SHEET in wb.sheetnames:
        del wb[MASK_SHEET]  # mantém o documento final só com a aba visível
    ws = wb.active
    systems_by_row = _row_systems(ws)
    rows_to_delete = []
    if keep_systems is not None:
        rows_to_delete = [
            row
            for row, system in systems_by_row.items()
            if system in REPRODUCTIVE_SYSTEMS and system not in keep_systems
        ]

    colocados: set[str] = set()
    ignorados: set[str] = set()
    sem_dados: list[str] = []

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        nome = ws.cell(row, ITEM_COL).value
        if nome is None:
            continue
        nome = str(nome).strip()
        if row in rows_to_delete:
            ignorados.add(nome)
            continue
        metrics = context.get(nome)
        if metrics is None:
            sem_dados.append(nome)
            continue
        aplicaveis = METRICS if mask is None else mask.get(row, METRICS)
        for metric in aplicaveis:
            col = METRIC_START_COL[metric]
            ws.cell(row, col).value = metrics[metric]["inicio"]
            ws.cell(row, col + 1).value = metrics[metric]["final"]
            ws.cell(row, col + 2).value = metrics[metric]["evolucao"]
        colocados.add(nome)

    _delete_rows_preserving_merges(ws, rows_to_delete)

    buffer = io.BytesIO()
    wb.save(buffer)
    return RenderResult(
        data=buffer.getvalue(),
        sem_dados=sem_dados,
        nao_colocados=sorted(set(context) - colocados - ignorados),
    )
